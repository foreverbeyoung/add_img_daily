# -*- coding:utf-8 -*-

import os
from math import ceil

import xlsxwriter
import requests
from fake_useragent import UserAgent
from openpyxl import Workbook,load_workbook

from openpyxl.drawing.image import Image

from hashlib import md5
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
from PIL import JpegImagePlugin
JpegImagePlugin._getmp = lambda  x:None

excel_dir = 'anker创新.xlsx'
img_org_save = r'F:\Amazon\看这里\test_imgs'

def get_header(referer="https://www.amazon.com/dp/product-reviews/B01MQU5LW7"):
    return {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'en,zh-CN;q=0.9,zh;q=0.8',
        'User-Agent': UserAgent().firefox,
        # 'User-Agent': random.choice(agent),
        'referer': referer,
    }
def seed_get():
    # return os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir, "demo.xlsx"))
    return os.path.abspath(os.path.join(os.path.dirname(__file__), excel_dir))

def file_name_encrypt(url):
    md5_inst = md5()
    md5_inst.update(url.encode())
    print(url, md5_inst.hexdigest())
    return md5_inst.hexdigest()

def save_org_img(url,filedir):
    if not os.path.exists(filedir):
        os.mkdir(filedir)
    try:
        res = requests.get(url,headers=get_header(),timeout=120)
        file_name = file_name_encrypt(url)
        img_fp= res.content
        with open(os.path.join(filedir,file_name)+'.jpeg', 'wb') as f:
            # print(os.path.join(filedir,file_name))
             # for data in res.iter_content(128):
             #    f.write(data)
            f.write(img_fp)
        f.close()
        # yield dir_position+file_name
        return os.path.join(filedir,file_name)+'.jpeg'
    except:
        return None

def img_position_join(x,y):
    if isinstance(y,int):
        y = get_column_letter(y)

    return y+str(x)


# def insert_img(sh, line, filedir,img_positon):
def insert_img(sh,line, filedir,img_positon):
    print(img_positon)
    sh.row_dimensions[line].height=50
    img = Image(filedir)
    print(img.width,img.height,img.format) #format 图片格式 quality图片质量
    print(filedir)
    img.format='jpeg'
    img.width = ceil(img.width * 55/ img.height)
    img.height = 55
    print(img.width,img.height)


    sh.add_image(img, img_positon)

def pd_read_excel():
    wb = Workbook()
    sh = wb.active
    df = pd.read_excel('demo.xlsx', sheet_name=0)
    col_name = df.columns.values.tolist()
    print(col_name)
    img_url_col_num = col_name.index('main_img_url') #返回img的列数
    max_column = len(col_name) #本表最大列数
    print(max_column)
    for idx, row in df.iterrows():
        print(row['main_img_url'])

        # d_row = {}
        # for column in col_name:
        #     d_row[column] = row[column]
        #     if column == 'output_date':
        #         s = str(row[column])
        #         d_row[column] = s[:4] + '-' + s[4:6] + '-' + s[6:]
        #     else:
        #         d_row[column] = row[column]
# pd_read_excel()

    #
if __name__ == "__main__":
    # wb = Workbook()
    # sh = wb.active
    wb = load_workbook(seed_get())
    ws = wb['Sheet1']

    print(ws.max_column)
    ws.insert_cols(2,1)  #插入第二列
    ws['B1']='img_exhibit'
    ws.column_dimensions['B'].width = 10  #设置列的宽度

#     for fv in ws[2:ws.max_row]:
    for fv in ws.iter_rows(min_row=2,min_col=1):
        if fv[2].value:
            img_dir = save_org_img(fv[2].value, img_org_save)
            print(img_dir)
            # input()
            if img_dir:
                insert_img(ws,fv[2].row,img_dir,img_position_join(fv[2].row,2))
    wb.save(excel_dir)

#         # for sv in fv:
# #         #     print(sv.value)
#         print(fv[2].value)
#         img_dir = save_org_img(fv[2].value,r'test_img')
#         print(img_dir)

#     print(ws.max_column)
#     print(get_column_letter(ws.max_column))
#     wb.save('demo.xlsx')

    # # sh.column_dimensions[IMG_COL].width= 10
    # sh.column_dimensions['AI'].width= 10
    # url = 'https://m.media-amazon.com/images/I/51tAycNDRfL._AC_SX679_.jpg'
    # file_name = save_org_img(url,r'F:/图片收集/')
    # insert_img(sh, 1, file_name)
    # wb.save('demo.xlsx')

# wb = openpyxl.load_workbook('demo.xlsx')
# # sheet = wb.get_sheet_by_name('Sheet')
# sheet = wb['Sheet']
# print(sheet.max_row)
# print(get_column_letter(sheet.max_column))
# print(sheet.max_column)

# def run():
#     df = pd.read_excel('demo.xlsx',sheet_name=0)
#     col_name = df.columns.values.tolist()
#     img_url_col_num = col_name.index('main_img_url') #返回img的列数
#     max_column = len(col_name) #本表最大列数
#     # print(max_column)
#     for idx, row in df.iterrows():
#         d_row = {}
#         for column in col_name:
#             d_row[column] = row[column]
#             if column == 'output_date':
#                 s = str(row[column])
#                 d_row[column] = s[:4] + '-' + s[4:6] + '-' + s[6:]
#             else:
#                 d_row[column] = row[column]
#
#
# run()